from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.contrib import messages
from django.contrib.auth.models import User
from django.db.models import F, Value, CharField, Count, Sum, Avg, Q, DecimalField, ExpressionWrapper
from django.db.models.functions import Coalesce, TruncMonth, Trunc
from django.http import HttpResponse
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.views.decorators.http import require_http_methods
from django.core.paginator import Paginator
from itertools import chain
from .models import Task, Project, CustomUser, Material, Equipment, MaterialCategory, MaterialSubcategory, MaterialTransaction, Warehouse, Subcontractor, EquipmentMaintenance, Milestone, BudgetItem, ResourceAllocation, TaskDependency, CostItem, CostCenter
from .forms import TaskForm, ProjectForm, MaterialForm, EquipmentForm, EquipmentUsageForm, EquipmentTransferForm, WarehouseForm, StockTransferForm, MaterialTransactionForm, SubcontractorForm, MaintenanceForm, MilestoneForm, BudgetItemForm, ResourceAllocationForm, TaskDependencyForm, CostItemForm
from .decorators import role_required, admin_required, project_manager_required, worker_or_above_required
from .roles import assign_role_to_user
import csv
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from decimal import Decimal
import json
from django.core.serializers.json import DjangoJSONEncoder
from .audit_models import AuditLog

# Create your views here.

@login_required
@role_required(['Admin', 'Project Manager', 'Worker', 'Viewer'])
def task_list(request):
    """View for listing tasks based on user's role"""
    user = request.user.customuser
    
    if user.has_role('Admin') or user.has_role('Project Manager'):
        # Admins and PMs can see all tasks
        tasks = Task.objects.all()
    elif user.has_role('Worker'):
        # Workers can only see their assigned tasks
        tasks = Task.objects.filter(assigned_to=user)
    else:  # Viewer
        # Viewers can see all tasks but with limited details
        tasks = Task.objects.all()
    
    return render(request, 'myapp/task_list.html', {'tasks': tasks})

@login_required
@role_required(['Admin', 'Project Manager'])
def task_create(request):
    """View for creating new tasks"""
    if request.method == 'POST':
        form = TaskForm(request.POST)
        if form.is_valid():
            task = form.save(commit=False)
            task.created_by = request.user.customuser
            task.save()
            task.log_action('CREATE', user=request.user.customuser, request=request)
            messages.success(request, 'Task created successfully.')
            return redirect('task_list')
    else:
        form = TaskForm()
    return render(request, 'myapp/task_form.html', {'form': form})

@login_required
@role_required(['Admin', 'Project Manager', 'Worker'])
def task_update(request, pk):
    """View for updating tasks"""
    task = get_object_or_404(Task, pk=pk)
    user = request.user.customuser
    
    # Workers can only update their assigned tasks
    if user.has_role('Worker') and task.assigned_to != user:
        raise PermissionDenied
    
    if request.method == 'POST':
        old_task = Task.objects.get(pk=pk)  # Get old state for audit
        form = TaskForm(request.POST, instance=task)
        if form.is_valid():
            task = form.save()
            task.log_action('UPDATE', user=request.user.customuser, request=request, 
                          changes=task._get_changes(old_task))
            messages.success(request, 'Task updated successfully.')
            return redirect('task_list')
    else:
        form = TaskForm(instance=task)
    return render(request, 'myapp/task_form.html', {'form': form})

@login_required
@role_required(['Admin', 'Project Manager'])
def task_delete(request, pk):
    """View for deleting tasks"""
    task = get_object_or_404(Task, pk=pk)
    if request.method == 'POST':
        task.log_action('DELETE', user=request.user.customuser, request=request)
        task.delete()
        messages.success(request, 'Task deleted successfully.')
        return redirect('task_list')
    return render(request, 'myapp/task_confirm_delete.html', {'task': task})

@login_required
@admin_required
def user_management(request):
    """Admin view for managing users"""
    users = CustomUser.objects.all()
    return render(request, 'myapp/user_management.html', {'users': users})

@login_required
@project_manager_required
def project_dashboard(request):
    """Project manager dashboard"""
    projects = Project.objects.all()
    return render(request, 'myapp/project_dashboard.html', {'projects': projects})

@login_required
@worker_or_above_required
def material_list(request):
    """View for listing materials"""
    materials = Material.objects.all()
    return render(request, 'myapp/material_list.html', {'materials': materials})

@login_required
@worker_or_above_required
def equipment_list(request):
    """View for listing equipment"""
    equipment = Equipment.objects.all()
    return render(request, 'myapp/equipment_list.html', {'equipment': equipment})

@login_required
def home(request):
    """Home page view showing dashboard with recent activities."""
    # Get recent tasks with related data
    recent_tasks = Task.objects.select_related(
        'project', 'assigned_to'
    ).order_by('-created_at')[:5]

    # Get recent transactions with related data
    recent_transactions = MaterialTransaction.objects.select_related(
        'project', 'material'
    ).order_by('-created_at')[:5]

    # Get low stock materials
    low_stock_materials = Material.objects.filter(
        current_stock__lte=F('minimum_stock')
    ).annotate(
        stock_status=Value('Low', output_field=CharField())
    )[:5]

    context = {
        'recent_tasks': recent_tasks,
        'recent_transactions': recent_transactions,
        'low_stock_materials': low_stock_materials,
        'total_tasks': Task.objects.count(),
        'total_materials': Material.objects.count(),
        'total_projects': Project.objects.count()
    }
    return render(request, 'myapp/home.html', context)

@login_required
def project_list(request):
    projects = Project.objects.all()
    return render(request, 'myapp/project_list.html', {'projects': projects})

@login_required
def project_create(request):
    if request.method == 'POST':
        form = ProjectForm(request.POST)
        if form.is_valid():
            project = form.save(commit=False)
            project.created_by = request.user.customuser
            project.save()
            project.log_action('CREATE', user=request.user.customuser, request=request)
            messages.success(request, 'Project created successfully.')
            return redirect('myapp:project_list')
    else:
        form = ProjectForm()
    return render(request, 'myapp/project_form.html', {'form': form, 'action': 'Create'})

@login_required
def project_edit(request, pk):
    project = get_object_or_404(Project, pk=pk)
    if request.method == 'POST':
        old_project = Project.objects.get(pk=pk)  # Get old state for audit
        form = ProjectForm(request.POST, instance=project)
        if form.is_valid():
            project = form.save()
            project.log_action('UPDATE', user=request.user.customuser, request=request,
                             changes=project._get_changes(old_project))
            messages.success(request, 'Project updated successfully.')
            return redirect('myapp:project_list')
    else:
        form = ProjectForm(instance=project)
    return render(request, 'myapp/project_form.html', {'form': form, 'action': 'Edit'})

@login_required
def project_delete(request, pk):
    project = get_object_or_404(Project, pk=pk)
    if request.method == 'POST':
        project.log_action('DELETE', user=request.user.customuser, request=request)
        project.delete()
        messages.success(request, 'Project deleted successfully.')
        return redirect('myapp:project_list')
    return render(request, 'myapp/project_confirm_delete.html', {'project': project})

@login_required
def material_list(request):
    # Get all materials with their categories and subcategories
    materials = Material.objects.select_related('category', 'subcategory').all()
    
    # Handle search
    search_query = request.GET.get('search')
    if search_query:
        materials = materials.filter(
            models.Q(name__icontains=search_query) |
            models.Q(code__icontains=search_query) |
            models.Q(category__name__icontains=search_query) |
            models.Q(subcategory__name__icontains=search_query)
        )
    
    # Handle category filtering
    category_id = request.GET.get('category')
    if category_id:
        materials = materials.filter(category_id=category_id)
    
    # Handle subcategory filtering
    subcategory_id = request.GET.get('subcategory')
    if subcategory_id:
        materials = materials.filter(subcategory_id=subcategory_id)
    
    # Handle stock level filtering
    stock_filter = request.GET.get('stock')
    if stock_filter == 'low':
        materials = materials.filter(current_stock__lte=F('minimum_stock'))
    elif stock_filter == 'out':
        materials = materials.filter(current_stock=0)
    
    # Get categories and subcategories for the filter dropdowns
    categories = MaterialCategory.objects.all()
    subcategories = MaterialSubcategory.objects.all()
    
    context = {
        'materials': materials,
        'categories': categories,
        'subcategories': subcategories,
        'search_query': search_query,
        'selected_category': category_id,
        'selected_subcategory': subcategory_id,
        'stock_filter': stock_filter
    }
    
    return render(request, 'myapp/material_list.html', context)

@login_required
def material_create(request):
    if request.method == 'POST':
        form = MaterialForm(request.POST)
        if form.is_valid():
            material = form.save()
            messages.success(request, 'Material created successfully.')
            return redirect('myapp:material_list')
    else:
        form = MaterialForm()
    
    categories = MaterialCategory.objects.all()
    subcategories = MaterialSubcategory.objects.all()
    
    return render(request, 'myapp/create_material.html', {
        'form': form,
        'categories': categories,
        'subcategories': subcategories,
    })

@login_required
def material_edit(request, material_id):
    material = get_object_or_404(Material, id=material_id)
    if request.method == 'POST':
        old_material = Material.objects.get(id=material_id)  # Get old state for audit
        form = MaterialForm(request.POST, instance=material)
        if form.is_valid():
            material = form.save()
            material.log_action('UPDATE', user=request.user.customuser, request=request,
                              changes=material._get_changes(old_material))
            messages.success(request, 'Material updated successfully.')
            return redirect('material_detail', material_id=material.id)
    else:
        form = MaterialForm(instance=material)
    
    categories = MaterialCategory.objects.all()
    subcategories = MaterialSubcategory.objects.filter(category=material.category)
    
    return render(request, 'myapp/material_form.html', {
        'form': form,
        'categories': categories,
        'subcategories': subcategories,
    })

@login_required
def material_detail(request, material_id):
    material = get_object_or_404(Material.objects.select_related('category'), id=material_id)
    
    # Get recent transactions
    recent_deliveries = MaterialDelivery.objects.filter(material=material).order_by('-delivery_date')[:5]
    recent_consumptions = MaterialConsumption.objects.filter(material=material).order_by('-date')[:5]
    
    # Calculate statistics
    total_delivered = MaterialDelivery.objects.filter(material=material).aggregate(
        total=Coalesce(Sum('quantity'), Value(0))
    )['total']
    
    total_consumed = MaterialConsumption.objects.filter(material=material).aggregate(
        total=Coalesce(Sum('quantity'), Value(0))
    )['total']
    
    # Calculate average price from recent deliveries
    recent_prices = MaterialDelivery.objects.filter(material=material).order_by('-delivery_date')[:5]
    avg_price = recent_prices.aggregate(avg=models.Avg('unit_price'))['avg'] or material.unit_price
    
    # Get projects using this material
    projects_using = Project.objects.filter(
        id__in=MaterialConsumption.objects.filter(material=material).values_list('project_id', flat=True)
    ).distinct()
    
    context = {
        'material': material,
        'recent_deliveries': recent_deliveries,
        'recent_consumptions': recent_consumptions,
        'total_delivered': total_delivered,
        'total_consumed': total_consumed,
        'avg_price': avg_price,
        'projects_using': projects_using,
        'stock_status': 'Low' if material.current_stock <= material.minimum_stock else 'Normal',
        'stock_alert': material.current_stock <= material.minimum_stock
    }
    
    return render(request, 'myapp/materials/material_detail.html', context)

@login_required
def material_delete(request, material_id):
    material = get_object_or_404(Material, id=material_id)
    
    if request.method == 'POST':
        try:
            material_name = material.name
            material.log_action('DELETE', user=request.user.customuser, request=request)
            material.delete()
            messages.success(request, f'Material "{material_name}" deleted successfully!')
            return redirect('material_list')
        except Exception as e:
            messages.error(request, f'Error deleting material: {str(e)}')
            return redirect('material_detail', material_id=material_id)
    
    return render(request, 'myapp/materials/material_confirm_delete.html', {'material': material})

@login_required
def record_consumption(request):
    if request.method == 'POST':
        try:
            material_id = request.POST.get('material')
            project_id = request.POST.get('project')
            quantity = request.POST.get('quantity')
            unit_price = request.POST.get('unit_price')
            date = request.POST.get('date')
            notes = request.POST.get('notes')
            
            # Get the material
            material = Material.objects.get(id=material_id)
            
            # Check if we have enough stock
            if material.current_stock < Decimal(quantity):
                messages.error(request, f'Not enough stock! Available: {material.current_stock} {material.unit}')
                return redirect('record_consumption')
            
            # Create consumption record
            consumption = MaterialConsumption.objects.create(
                material_id=material_id,
                project_id=project_id,
                quantity=quantity,
                unit_price=unit_price,
                date=date,
                notes=notes,
                recorded_by=request.user.customuser
            )
            
            # Update material stock
            material.current_stock -= Decimal(quantity)
            material.save()
            
            messages.success(request, 'Consumption recorded successfully!')
            return redirect('material_detail', material_id=material_id)
            
        except Exception as e:
            messages.error(request, f'Error recording consumption: {str(e)}')
            return redirect('record_consumption')
    
    # Get active projects and materials for the form
    projects = Project.objects.filter(status='ONGOING')
    materials = Material.objects.all()
    
    return render(request, 'myapp/materials/consumption_form.html', {
        'projects': projects,
        'materials': materials
    })

@login_required
def record_delivery(request):
    if request.method == 'POST':
        try:
            material_id = request.POST.get('material')
            project_id = request.POST.get('project')
            quantity = request.POST.get('quantity')
            unit_price = request.POST.get('unit_price')
            supplier = request.POST.get('supplier')
            delivery_date = request.POST.get('delivery_date')
            invoice_number = request.POST.get('invoice_number')
            notes = request.POST.get('notes')
            
            # Create delivery record
            delivery = MaterialDelivery.objects.create(
                material_id=material_id,
                project_id=project_id,
                quantity=quantity,
                unit_price=unit_price,
                supplier=supplier,
                delivery_date=delivery_date,
                invoice_number=invoice_number,
                notes=notes,
                received_by=request.user.customuser
            )
            
            # Update material stock and price
            material = Material.objects.get(id=material_id)
            material.current_stock += Decimal(quantity)
            material.unit_price = unit_price  # Update to latest price
            material.save()
            
            messages.success(request, 'Delivery recorded successfully!')
            return redirect('material_detail', material_id=material_id)
            
        except Exception as e:
            messages.error(request, f'Error recording delivery: {str(e)}')
            return redirect('record_delivery')
    
    # Get active projects and materials for the form
    projects = Project.objects.filter(status='ONGOING')
    materials = Material.objects.all()
    
    return render(request, 'myapp/materials/delivery_form.html', {
        'projects': projects,
        'materials': materials
    })

# Category Management Views
@login_required
def category_list(request):
    categories = MaterialCategory.objects.all()
    return render(request, 'myapp/manage_categories.html', {
        'categories': categories,
        'main_categories': MaterialCategory.objects.filter(parent=None)
    })

@login_required
def category_create(request):
    if request.method == 'POST':
        try:
            name = request.POST.get('name')
            parent_id = request.POST.get('parent')
            
            if not name:
                return JsonResponse({'success': False, 'error': 'Category name is required'})
            
            category = MaterialCategory.objects.create(
                name=name,
                parent_id=parent_id if parent_id else None
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Category "{category.name}" created successfully!'
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

@login_required
def category_edit(request, pk):
    try:
        category = get_object_or_404(MaterialCategory, id=pk)
        
        if request.method == 'POST':
            name = request.POST.get('name')
            parent_id = request.POST.get('parent')
            
            if not name:
                return JsonResponse({'success': False, 'error': 'Category name is required'})
            
            # Prevent circular parent reference
            if parent_id and int(parent_id) == category.id:
                return JsonResponse({'success': False, 'error': 'A category cannot be its own parent'})
            
            category.name = name
            category.parent_id = parent_id if parent_id else None
            category.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Category "{category.name}" updated successfully!'
            })
        
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
        
    except MaterialCategory.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Category not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def category_delete(request, pk):
    try:
        category = get_object_or_404(MaterialCategory, id=pk)
        
        if request.method == 'POST':
            category_name = category.name
            
            # Check if category has materials
            if category.material_set.exists():
                return JsonResponse({
                    'success': False,
                    'error': 'Cannot delete category that has materials assigned to it'
                })
            
            # Check if category has subcategories
            if category.materialcategory_set.exists():
                return JsonResponse({
                    'success': False,
                    'error': 'Cannot delete category that has subcategories'
                })
            
            category.log_action('DELETE', user=request.user.customuser, request=request)
            category.delete()
            return JsonResponse({
                'success': True,
                'message': f'Category "{category_name}" deleted successfully!'
            })
        
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
        
    except MaterialCategory.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Category not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def get_subcategories(request, category_id):
    subcategories = MaterialSubcategory.objects.filter(category_id=category_id)
    data = [{'id': sub.id, 'name': sub.name} for sub in subcategories]
    return JsonResponse(data, safe=False)

# API endpoint for dynamic subcategory loading
@login_required
def get_subcategories(request, category_id):
    subcategories = MaterialSubcategory.objects.filter(category_id=category_id).values('id', 'name')
    return JsonResponse(list(subcategories), safe=False)

# Subcategory Management Views
@login_required
def subcategory_create(request):
    if request.method == 'POST':
        try:
            name = request.POST.get('name')
            category_id = request.POST.get('category')
            
            if not name:
                return JsonResponse({'success': False, 'error': 'Subcategory name is required'})
            if not category_id:
                return JsonResponse({'success': False, 'error': 'Parent category is required'})
            
            category = get_object_or_404(MaterialCategory, id=category_id)
            subcategory = MaterialSubcategory.objects.create(
                name=name,
                category=category
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Subcategory "{subcategory.name}" created successfully!'
            })
            
        except MaterialCategory.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Parent category not found'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

@login_required
def subcategory_edit(request, pk):
    try:
        subcategory = get_object_or_404(MaterialSubcategory, id=pk)
        
        if request.method == 'POST':
            name = request.POST.get('name')
            category_id = request.POST.get('category')
            
            if not name:
                return JsonResponse({'success': False, 'error': 'Subcategory name is required'})
            if not category_id:
                return JsonResponse({'success': False, 'error': 'Parent category is required'})
            
            category = get_object_or_404(MaterialCategory, id=category_id)
            
            subcategory.name = name
            subcategory.category = category
            subcategory.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Subcategory "{subcategory.name}" updated successfully!'
            })
        
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
        
    except MaterialSubcategory.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Subcategory not found'})
    except MaterialCategory.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Parent category not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def subcategory_delete(request, pk):
    try:
        subcategory = get_object_or_404(MaterialSubcategory, id=pk)
        
        if request.method == 'POST':
            subcategory_name = subcategory.name
            
            # Check if subcategory has materials
            if subcategory.material_set.exists():
                return JsonResponse({
                    'success': False,
                    'error': 'Cannot delete subcategory that has materials assigned to it'
                })
            
            subcategory.log_action('DELETE', user=request.user.customuser, request=request)
            subcategory.delete()
            return JsonResponse({
                'success': True,
                'message': f'Subcategory "{subcategory_name}" deleted successfully!'
            })
        
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
        
    except MaterialSubcategory.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Subcategory not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def get_subcategories(request, category_id):
    try:
        subcategories = MaterialSubcategory.objects.filter(category_id=category_id)
        data = [{'id': sub.id, 'name': sub.name} for sub in subcategories]
        return JsonResponse({'success': True, 'subcategories': data})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def all_projects(request):
    projects = Project.objects.select_related('manager')
    
    # Handle search
    search_query = request.GET.get('search')
    if search_query:
        projects = projects.filter(
            models.Q(name__icontains=search_query) |
            models.Q(code__icontains=search_query) |
            models.Q(location__icontains=search_query)
        )
    
    # Handle status filtering
    status = request.GET.get('status')
    if status:
        projects = projects.filter(status=status.upper())
    
    # Handle sorting
    sort_by = request.GET.get('sort')
    if sort_by:
        if sort_by == 'deadline':
            projects = projects.order_by('end_date')
        elif sort_by == 'budget':
            projects = projects.order_by('-total_budget')
        elif sort_by == 'name':
            projects = projects.order_by('name')
        elif sort_by == 'progress':
            projects = projects.order_by('-progress')
    else:
        # Default sort by creation date
        projects = projects.order_by('-created_at')
    
    # Calculate progress for each project
    for project in projects:
        # You can customize this calculation based on your needs
        if project.status == 'COMPLETED':
            project.progress = 100
        elif project.status == 'NOT_STARTED':
            project.progress = 0
        else:
            # Calculate based on timeline
            total_days = (project.end_date - project.start_date).days
            days_passed = (timezone.now().date() - project.start_date).days
            project.progress = min(100, max(0, int((days_passed / total_days) * 100)))
        
        # Add status color for badges
        project.status_color = {
            'NOT_STARTED': 'info',
            'ONGOING': 'primary',
            'COMPLETED': 'success',
            'ON_HOLD': 'warning',
            'CANCELLED': 'danger'
        }.get(project.status, 'secondary')
    
    return render(request, 'myapp/project_list.html', {'projects': projects})

@login_required
def create_project(request):
    if request.method == 'POST':
        # Get form data
        name = request.POST.get('name')
        code = request.POST.get('code')
        location = request.POST.get('location')
        description = request.POST.get('description')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        total_budget = request.POST.get('total_budget')
        status = request.POST.get('status')
        
        try:
            # Create new project
            project = Project.objects.create(
                name=name,
                code=code,
                location=location,
                description=description,
                start_date=start_date,
                end_date=end_date,
                total_budget=total_budget,
                status=status,
                manager=request.user
            )
            project.log_action('CREATE', user=request.user.customuser, request=request)
            messages.success(request, 'Project created successfully!')
            return redirect('myapp:project_detail', project_id=project.id)
        except Exception as e:
            messages.error(request, f'Error creating project: {str(e)}')
            return redirect('myapp:create_project')
    
    return render(request, 'myapp/create_project.html')

@login_required
def project_detail(request, project_id):
    project = get_object_or_404(Project.objects.select_related('manager'), id=project_id)
    
    # Calculate days remaining
    today = timezone.now().date()
    days_remaining = (project.end_date - today).days if project.end_date > today else 0
    
    # Calculate budget used (you'll need to implement actual budget tracking)
    budget_used = 45  # Example value, replace with actual calculation
    
    # Get material count for this project
    material_count = MaterialConsumption.objects.filter(project=project).count()
    
    # Example data for team members (implement your own team member model)
    team_members_count = 5  # Replace with actual count
    
    # Example data for documents (implement your own document model)
    document_count = 3  # Replace with actual count
    
    # Example data for issues (implement your own issue tracking)
    open_issues_count = 2  # Replace with actual count
    
    # Example data for tasks (implement your own task model)
    completed_tasks = 8  # Replace with actual count
    total_tasks = 12  # Replace with actual count
    
    # Get recent activities (implement your own activity tracking)
    recent_activities = [
        {
            'title': 'Material Delivered',
            'description': '20 tons of cement delivered to site',
            'timestamp': timezone.now() - timedelta(hours=2)
        },
        {
            'title': 'Task Completed',
            'description': 'Foundation work completed',
            'timestamp': timezone.now() - timedelta(days=1)
        },
        {
            'title': 'New Team Member',
            'description': 'John Doe joined the project',
            'timestamp': timezone.now() - timedelta(days=2)
        }
    ]  # Replace with actual activities
    
    # Get project milestones
    milestones = project.milestones.all().order_by('due_date')[:5]
    total_milestones = project.milestones.count()
    completed_milestones = project.milestones.filter(is_completed=True).count()
    
    context = {
        'project': project,
        'days_remaining': days_remaining,
        'budget_used': budget_used,
        'material_count': material_count,
        'team_members_count': team_members_count,
        'document_count': document_count,
        'open_issues_count': open_issues_count,
        'completed_tasks': completed_tasks,
        'total_tasks': total_tasks,
        'recent_activities': recent_activities,
        'milestones': milestones,
        'total_milestones': total_milestones,
        'completed_milestones': completed_milestones
    }
    
    return render(request, 'myapp/project_detail.html', context)

@login_required
def edit_project(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    
    if request.method == 'POST':
        # Get form data
        project.name = request.POST.get('name')
        project.code = request.POST.get('code')
        project.location = request.POST.get('location')
        project.description = request.POST.get('description')
        project.start_date = request.POST.get('start_date')
        project.end_date = request.POST.get('end_date')
        project.total_budget = request.POST.get('total_budget')
        project.status = request.POST.get('status')
        
        try:
            project.save()
            project.log_action('UPDATE', user=request.user.customuser, request=request)
            messages.success(request, 'Project updated successfully.')
            return redirect('myapp:project_detail', project_id=project.id)
        except Exception as e:
            messages.error(request, f'Error updating project: {str(e)}')
    
    return render(request, 'myapp/edit_project.html', {'project': project})

@login_required
def project_dashboard(request):
    """Dashboard view showing project statistics and recent activities."""
    # Get all projects
    total_projects = Project.objects.count()
    active_projects = Project.objects.filter(status='ACTIVE').count()
    
    # Get projects by status
    projects_by_status = Project.objects.values('status').annotate(
        count=Count('id')
    ).order_by('status')
    
    # Get recent tasks
    recent_tasks = Task.objects.select_related(
        'project', 'assigned_to'
    ).order_by('-created_at')[:10]
    
    # Get tasks by status
    tasks_by_status = Task.objects.values('status').annotate(
        count=Count('id')
    ).order_by('status')
    
    # Get material consumption by project
    material_by_project = MaterialTransaction.objects.filter(
        transaction_type='CONSUMPTION'
    ).values('project__name').annotate(
        total_consumption=Count('id')
    ).order_by('-total_consumption')[:5]
    
    context = {
        'total_projects': total_projects,
        'active_projects': active_projects,
        'projects_by_status': projects_by_status,
        'recent_tasks': recent_tasks,
        'tasks_by_status': tasks_by_status,
        'material_by_project': material_by_project,
    }
    return render(request, 'myapp/project_dashboard.html', context)

@login_required
def add_project(request):
    if request.method == 'POST':
        form = ProjectForm(request.POST)
        if form.is_valid():
            project = form.save(commit=False)
            project.created_by = request.user
            project.save()
            project.log_action('CREATE', user=request.user.customuser, request=request)
            messages.success(request, 'Project created successfully.')
            return redirect('myapp:project_list')
    else:
        form = ProjectForm()
    return render(request, 'myapp/add_project.html', {'form': form})

@login_required
def edit_project(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    if request.method == 'POST':
        form = ProjectForm(request.POST, instance=project)
        if form.is_valid():
            form.save()
            project.log_action('UPDATE', user=request.user.customuser, request=request)
            messages.success(request, 'Project updated successfully.')
            return redirect('myapp:project_list')
    else:
        form = ProjectForm(instance=project)
    return render(request, 'myapp/edit_project.html', {'form': form, 'project': project})

@login_required
def task_dashboard(request):
    """Dashboard view showing task statistics and recent tasks."""
    # Get task statistics
    total_tasks = Task.objects.count()
    tasks_by_status = Task.objects.values('status').annotate(
        count=Count('id')
    ).order_by('status')
    
    tasks_by_priority = Task.objects.values('priority').annotate(
        count=Count('id')
    ).order_by('priority')
    
    tasks_by_assignee = Task.objects.values(
        'assigned_to__user__username'
    ).annotate(
        count=Count('id')
    ).order_by('-count')[:5]
    
    recent_tasks = Task.objects.select_related(
        'project', 'assigned_to', 'assigned_to__user'
    ).order_by('-created_at')[:10]
    
    overdue_tasks = Task.objects.filter(
        due_date__lt=timezone.now().date(),
        status__in=['PENDING', 'IN_PROGRESS']
    ).count()

    context = {
        'total_tasks': total_tasks,
        'tasks_by_status': tasks_by_status,
        'tasks_by_priority': tasks_by_priority,
        'tasks_by_assignee': tasks_by_assignee,
        'recent_tasks': recent_tasks,
        'overdue_tasks': overdue_tasks
    }
    return render(request, 'myapp/task_dashboard.html', context)

@login_required
def task_list(request, project_id=None):
    """View for listing tasks, optionally filtered by project."""
    if project_id:
        project = get_object_or_404(Project, id=project_id)
        tasks = Task.objects.filter(project=project)
    else:
        tasks = Task.objects.all()
    
    tasks = tasks.select_related('project', 'assigned_to').order_by('-created_at')
    return render(request, 'myapp/task_list.html', {'tasks': tasks})

@login_required
def add_task(request):
    if request.method == 'POST':
        form = TaskForm(request.POST)
        if form.is_valid():
            task = form.save(commit=False)
            task.created_by = request.user
            task.save()
            task.log_action('CREATE', user=request.user.customuser, request=request)
            messages.success(request, 'Task created successfully.')
            return redirect('myapp:task_list', project_id=task.project.id)
    else:
        form = TaskForm()
    return render(request, 'myapp/add_task.html', {'form': form})

@login_required
def edit_task(request, task_id):
    task = get_object_or_404(Task, id=task_id)
    if request.method == 'POST':
        form = TaskForm(request.POST, instance=task)
        if form.is_valid():
            form.save()
            task.log_action('UPDATE', user=request.user.customuser, request=request)
            messages.success(request, 'Task updated successfully.')
            return redirect('myapp:task_list', project_id=task.project.id)
    else:
        form = TaskForm(instance=task)
    return render(request, 'myapp/edit_task.html', {'form': form, 'task': task})

@login_required
def task_create(request):
    """Create a new task."""
    if request.method == 'POST':
        form = TaskForm(request.POST)
        if form.is_valid():
            task = form.save(commit=False)
            task.created_by = request.user
            task.save()
            task.log_action('CREATE', user=request.user.customuser, request=request)
            messages.success(request, 'Task created successfully.')
            return redirect('myapp:task_list')
    else:
        form = TaskForm()
    
    return render(request, 'myapp/task_form.html', {
        'form': form,
        'title': 'Create Task'
    })

@login_required
def task_edit(request, pk):
    """Edit an existing task."""
    task = get_object_or_404(Task, pk=pk)
    if request.method == 'POST':
        form = TaskForm(request.POST, instance=task)
        if form.is_valid():
            task = form.save()
            task.log_action('UPDATE', user=request.user.customuser, request=request)
            messages.success(request, 'Task updated successfully.')
            return redirect('myapp:task_list')
    else:
        form = TaskForm(instance=task)
    
    return render(request, 'myapp/task_form.html', {
        'form': form,
        'title': 'Edit Task',
        'task': task
    })

@login_required
def task_delete(request, pk):
    task = get_object_or_404(Task, pk=pk)
    if request.method == 'POST':
        task.log_action('DELETE', user=request.user.customuser, request=request)
        task.delete()
        messages.success(request, 'Task deleted successfully.')
        return redirect('myapp:task_list')
    return render(request, 'myapp/task_confirm_delete.html', {'task': task})

@login_required
def task_import(request):
    if request.method == 'POST' and request.FILES.get('task_file'):
        file = request.FILES['task_file']
        filename = file.name.lower()
        tasks_created = 0
        errors = []

        try:
            if filename.endswith('.csv'):
                # Handle CSV file
                decoded_file = file.read().decode('utf-8').splitlines()
                reader = csv.DictReader(decoded_file)
                data = list(reader)
            elif filename.endswith(('.xlsx', '.xls')):
                # Handle Excel file
                wb = openpyxl.load_workbook(file)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                data = []
                for row in ws.iter_rows(min_row=2):
                    row_data = {}
                    for header, cell in zip(headers, row):
                        row_data[header] = cell.value
                    data.append(row_data)
            else:
                messages.error(request, 'Unsupported file format. Please upload a CSV or Excel file.')
                return redirect('myapp:task_import')

            # Process the data
            for row in data:
                try:
                    # Get project by name
                    try:
                        project = Project.objects.get(name=row['project'])
                    except Project.DoesNotExist:
                        raise ValidationError(f"Project '{row['project']}' not found")

                    # Get user by username
                    try:
                        assigned_to = User.objects.get(username=row['assigned_to'])
                    except User.DoesNotExist:
                        raise ValidationError(f"User '{row['assigned_to']}' not found")

                    # Parse dates in DD-MM-YYYY format
                    start_date = datetime.strptime(str(row['start_date']), '%d-%m-%Y').date() if row.get('start_date') else None
                    due_date = datetime.strptime(str(row['due_date']), '%d-%m-%Y').date() if row.get('due_date') else None

                    # Create task
                    task = Task(
                        title=row['title'],
                        description=row.get('description', ''),
                        project=project,
                        assigned_to=assigned_to,
                        status=row.get('status', 'NOT_STARTED'),
                        priority=row.get('priority', 'MEDIUM'),
                        progress=int(row.get('progress', 0)),
                        start_date=start_date,
                        due_date=due_date,
                        created_by=request.user
                    )
                    task.save()
                    task.log_action('CREATE', user=request.user.customuser, request=request)
                    tasks_created += 1
                except (KeyError, ValueError, ValidationError) as e:
                    errors.append(f"Error in row {tasks_created + 1}: {str(e)}")
                    continue

            if tasks_created > 0:
                success_msg = f'Successfully imported {tasks_created} tasks.'
                if errors:
                    success_msg += f' ({len(errors)} rows had errors)'
                messages.success(request, success_msg)
            if errors:
                messages.warning(request, 'Some rows had errors: ' + '; '.join(errors[:5]) + 
                               ('...' if len(errors) > 5 else ''))
            
        except Exception as e:
            messages.error(request, f'Error processing file: {str(e)}')
        
        return redirect('myapp:task_list')

    return render(request, 'myapp/task_import.html')

@login_required
def equipment_list(request):
    equipment = Equipment.objects.all()
    return render(request, 'myapp/equipment_list.html', {'equipment': equipment})

@login_required
def equipment_create(request):
    if request.method == 'POST':
        form = EquipmentForm(request.POST)
        if form.is_valid():
            equipment = form.save()
            equipment.log_action('CREATE', user=request.user.customuser, request=request)
            messages.success(request, 'Equipment created successfully.')
            return redirect('myapp:equipment_list')
    else:
        form = EquipmentForm()
    return render(request, 'myapp/equipment_form.html', {'form': form, 'action': 'Create'})

@login_required
def equipment_edit(request, pk):
    equipment = get_object_or_404(Equipment, pk=pk)
    if request.method == 'POST':
        form = EquipmentForm(request.POST, instance=equipment)
        if form.is_valid():
            form.save()
            equipment.log_action('UPDATE', user=request.user.customuser, request=request)
            messages.success(request, 'Equipment updated successfully.')
            return redirect('myapp:equipment_list')
    else:
        form = EquipmentForm(instance=equipment)
    return render(request, 'myapp/equipment_form.html', {'form': form, 'action': 'Edit', 'equipment': equipment})

@login_required
def equipment_delete(request, pk):
    equipment = get_object_or_404(Equipment, pk=pk)
    if request.method == 'POST':
        equipment.log_action('DELETE', user=request.user.customuser, request=request)
        equipment.delete()
        messages.success(request, 'Equipment deleted successfully.')
        return redirect('myapp:equipment_list')
    return render(request, 'myapp/equipment_confirm_delete.html', {'equipment': equipment})

@login_required
def equipment_import(request):
    if request.method == 'POST' and request.FILES.get('equipment_file'):
        file = request.FILES['equipment_file']
        try:
            decoded_file = file.read().decode('utf-8').splitlines()
            reader = csv.DictReader(decoded_file)
            
            for row in reader:
                equipment = Equipment(
                    name=row['name'],
                    code=row['code'],
                    description=row.get('description', ''),
                    equipment_type=row['equipment_type'],
                    manufacturer=row['manufacturer'],
                    model_number=row['model_number'],
                    serial_number=row['serial_number'],
                    purchase_date=datetime.strptime(row['purchase_date'], '%Y-%m-%d').date(),
                    purchase_cost=Decimal(row['purchase_cost']),
                    status=row.get('status', 'AVAILABLE'),
                    current_location=row['current_location']
                )
                equipment.save()
                equipment.log_action('CREATE', user=request.user.customuser, request=request)
            
            messages.success(request, 'Equipment imported successfully.')
            return redirect('myapp:equipment_list')
        except Exception as e:
            messages.error(request, f'Error importing equipment: {str(e)}')
    
    return render(request, 'myapp/equipment_import.html')

@login_required
def equipment_usage(request):
    if request.method == 'POST':
        form = EquipmentUsageForm(request.POST)
        if form.is_valid():
            usage = form.save(commit=False)
            usage.operator = request.user.customuser
            usage.save()
            usage.log_action('CREATE', user=request.user.customuser, request=request)
            messages.success(request, 'Equipment usage recorded successfully.')
            return redirect('myapp:equipment_list')
    else:
        form = EquipmentUsageForm()
    return render(request, 'myapp/equipment_usage_form.html', {'form': form})

@login_required
def equipment_transfer(request):
    if request.method == 'POST':
        form = EquipmentTransferForm(request.POST)
        if form.is_valid():
            transfer = form.save(commit=False)
            transfer.transferred_by = request.user.customuser
            transfer.save()
            transfer.log_action('CREATE', user=request.user.customuser, request=request)
            
            equipment = transfer.equipment
            equipment.current_location = transfer.to_location
            equipment.save()
            equipment.log_action('UPDATE', user=request.user.customuser, request=request)
            
            messages.success(request, 'Equipment transfer recorded successfully.')
            return redirect('myapp:equipment_list')
    else:
        form = EquipmentTransferForm()
    return render(request, 'myapp/equipment_transfer_form.html', {'form': form})

@login_required
def material_dashboard(request):
    """Dashboard view for material management."""
    # Get total materials and categories
    total_materials = Material.objects.count()
    total_categories = MaterialCategory.objects.count()
    
    # Get low stock materials
    low_stock_materials = Material.objects.filter(current_stock__lte=F('minimum_stock'))
    low_stock_count = low_stock_materials.count()
    
    # Get out of stock materials
    out_of_stock = Material.objects.filter(current_stock=0).count()
    
    # Get recent transactions
    recent_transactions = MaterialTransaction.objects.select_related(
        'material', 'project'
    ).order_by('-date')[:10]
    
    # Get materials by category
    categories = MaterialCategory.objects.annotate(
        material_count=Count('material')
    ).order_by('-material_count')[:5]
    
    context = {
        'total_materials': total_materials,
        'total_categories': total_categories,
        'low_stock_count': low_stock_count,
        'out_of_stock': out_of_stock,
        'low_stock_materials': low_stock_materials[:5],
        'recent_transactions': recent_transactions,
        'categories': categories,
    }
    return render(request, 'myapp/material_dashboard.html', context)

@login_required
def material_create(request):
    if request.method == 'POST':
        form = MaterialForm(request.POST)
        if form.is_valid():
            material = form.save()
            material.log_action('CREATE', user=request.user.customuser, request=request)
            messages.success(request, 'Material created successfully.')
            return redirect('myapp:material_list')
    else:
        form = MaterialForm()
    return render(request, 'myapp/create_material.html', {'form': form, 'action': 'Create'})

@login_required
def material_edit(request, pk):
    material = get_object_or_404(Material, pk=pk)
    if request.method == 'POST':
        form = MaterialForm(request.POST, instance=material)
        if form.is_valid():
            form.save()
            material.log_action('UPDATE', user=request.user.customuser, request=request)
            messages.success(request, 'Material updated successfully.')
            return redirect('myapp:material_list')
    else:
        form = MaterialForm(instance=material)
    return render(request, 'myapp/material_form.html', {'form': form, 'action': 'Edit', 'material': material})

@login_required
def material_delete(request, pk):
    material = get_object_or_404(Material, pk=pk)
    if request.method == 'POST':
        material.log_action('DELETE', user=request.user.customuser, request=request)
        material.delete()
        messages.success(request, 'Material deleted successfully.')
        return redirect('myapp:material_list')
    return render(request, 'myapp/material_confirm_delete.html', {'material': material})

@login_required
def material_import(request):
    if request.method == 'POST' and request.FILES.get('material_file'):
        file = request.FILES['material_file']
        try:
            decoded_file = file.read().decode('utf-8').splitlines()
            reader = csv.DictReader(decoded_file)
            
            for row in reader:
                material = Material(
                    name=row['name'],
                    code=row['code'],
                    category_id=row['category_id'],
                    subcategory_id=row.get('subcategory_id'),
                    description=row.get('description', ''),
                    unit=row['unit'],
                    unit_price=Decimal(row['unit_price']),
                    minimum_stock=Decimal(row.get('minimum_stock', '0')),
                    current_stock=Decimal(row.get('current_stock', '0'))
                )
                material.save()
                material.log_action('CREATE', user=request.user.customuser, request=request)
            
            messages.success(request, 'Materials imported successfully.')
            return redirect('myapp:material_list')
        except Exception as e:
            messages.error(request, f'Error importing materials: {str(e)}')
    
    return render(request, 'myapp/material_import.html')

@login_required
def inventory_dashboard(request):
    """Dashboard view for inventory management."""
    # Get total materials count
    total_materials = Material.objects.count()
    
    # Get low stock materials count
    low_stock_materials = Material.objects.filter(current_stock__lte=F('minimum_stock')).count()
    
    # Get recent transactions
    recent_transactions = MaterialTransaction.objects.select_related(
        'material', 'project'
    ).order_by('-date')[:10]
    
    # Get total stock value
    total_stock_value = Material.objects.aggregate(total_stock_value=Sum(F('current_stock') * F('unit_price')))['total_stock_value'] or 0
    
    # Get total stock quantity
    total_stock_quantity = Material.objects.aggregate(total_stock_quantity=Sum('current_stock'))['total_stock_quantity'] or 0
    
    # Get average stock level
    average_stock_level = Material.objects.aggregate(average_stock_level=Avg('current_stock'))['average_stock_level'] or 0
    
    context = {
        'total_materials': total_materials,
        'low_stock_materials': low_stock_materials,
        'recent_transactions': recent_transactions,
        'total_stock_value': total_stock_value,
        'total_stock_quantity': total_stock_quantity,
        'average_stock_level': average_stock_level,
    }
    return render(request, 'myapp/inventory/dashboard.html', context)

@login_required
def create_warehouse(request):
    if request.method == 'POST':
        form_data = request.POST.copy()
        form = WarehouseForm(form_data)
        if form.is_valid():
            warehouse = form.save(commit=False)
            warehouse.created_by = request.user.customuser
            warehouse.save()
            warehouse.log_action('CREATE', user=request.user.customuser, request=request)
            messages.success(request, 'Warehouse created successfully.')
            return redirect('myapp:warehouse_list')
    else:
        form = WarehouseForm()

    context = {
        'form': form,
        'projects': Project.objects.filter(status='ACTIVE'),
        'users': User.objects.filter(is_active=True),
        'action': 'Create'
    }
    return render(request, 'myapp/inventory/warehouse_form.html', context)

@login_required
def edit_warehouse(request, pk):
    warehouse = get_object_or_404(Warehouse, pk=pk)
    if request.method == 'POST':
        form = WarehouseForm(request.POST, instance=warehouse)
        if form.is_valid():
            form.save()
            warehouse.log_action('UPDATE', user=request.user.customuser, request=request)
            messages.success(request, 'Warehouse updated successfully.')
            return redirect('myapp:warehouse_list')
    else:
        form = WarehouseForm(instance=warehouse)

    context = {
        'form': form,
        'warehouse': warehouse,
        'projects': Project.objects.filter(status='ACTIVE'),
        'users': User.objects.filter(is_active=True),
        'action': 'Edit'
    }
    return render(request, 'myapp/inventory/warehouse_form.html', context)

@login_required
def warehouse_list(request):
    warehouses = Warehouse.objects.all()
    context = {
        'warehouses': warehouses,
    }
    return render(request, 'myapp/inventory/warehouse_list.html', context)

@login_required
def stock_transfer(request):
    if request.method == 'POST':
        form = StockTransferForm(request.POST)
        if form.is_valid():
            # First create the outgoing transaction
            outgoing = form.save(commit=False)
            outgoing.transaction_type = 'CONSUMPTION'
            outgoing.warehouse = form.cleaned_data['source_warehouse']
            outgoing.created_by = request.user.customuser
            outgoing.save()
            outgoing.log_action('CREATE', user=request.user.customuser, request=request)
            
            # Then create the incoming transaction
            incoming = MaterialTransaction(
                material=outgoing.material,
                quantity=outgoing.quantity,
                transaction_type='DELIVERY',
                warehouse=form.cleaned_data['destination_warehouse'],
                notes=outgoing.notes,
                created_by=request.user.customuser
            )
            incoming.save()
            incoming.log_action('CREATE', user=request.user.customuser, request=request)
            
            messages.success(request, 'Stock transfer recorded successfully.')
            return redirect('myapp:inventory_dashboard')
    else:
        form = StockTransferForm()
    
    return render(request, 'myapp/inventory/stock_transfer.html', {'form': form})

@login_required
def material_transaction(request):
    """View for listing material transactions with filtering."""
    # Get filter parameters
    material_id = request.GET.get('material')
    warehouse_id = request.GET.get('warehouse')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    # Base queryset
    transactions = MaterialTransaction.objects.select_related(
        'material', 'warehouse', 'project'
    ).order_by('-date')

    # Apply filters
    if material_id:
        transactions = transactions.filter(material_id=material_id)
    if warehouse_id:
        transactions = transactions.filter(warehouse_id=warehouse_id)
    if date_from:
        transactions = transactions.filter(date__gte=date_from)
    if date_to:
        transactions = transactions.filter(date__lte=date_to)

    # Separate transactions by type
    delivery_transactions = transactions.filter(transaction_type='delivery')
    consumption_transactions = transactions.filter(transaction_type='consumption')

    context = {
        'delivery_transactions': delivery_transactions,
        'consumption_transactions': consumption_transactions,
        'materials': Material.objects.all(),
        'warehouses': Warehouse.objects.all(),
        'selected_material': material_id,
        'selected_warehouse': warehouse_id,
        'date_from': date_from,
        'date_to': date_to,
    }
    
    return render(request, 'myapp/material_transaction.html', context)

@login_required
def record_delivery(request):
    """View for recording material deliveries."""
    if request.method == 'POST':
        material = get_object_or_404(Material, id=request.POST.get('material'))
        warehouse = get_object_or_404(Warehouse, id=request.POST.get('warehouse'))
        quantity = Decimal(request.POST.get('quantity'))
        unit_price = Decimal(request.POST.get('unit_price') or 0)

        # Create the transaction
        transaction = MaterialTransaction.objects.create(
            material=material,
            warehouse=warehouse,
            transaction_type='delivery',
            quantity=quantity,
            unit_price=unit_price,
            supplier=request.POST.get('supplier'),
            invoice_number=request.POST.get('invoice_number'),
            notes=request.POST.get('notes'),
        )
        transaction.log_action('CREATE', user=request.user.customuser, request=request)

        # Update material stock
        material.current_stock += quantity
        material.save()
        material.log_action('UPDATE', user=request.user.customuser, request=request)

        messages.success(request, 'Material delivery recorded successfully.')
        return redirect('myapp:material_transaction')

    context = {
        'materials': Material.objects.all(),
        'warehouses': Warehouse.objects.all(),
    }
    return render(request, 'myapp/record_delivery.html', context)

@login_required
def record_consumption(request):
    """View for recording material consumption."""
    if request.method == 'POST':
        material = get_object_or_404(Material, id=request.POST.get('material'))
        warehouse = get_object_or_404(Warehouse, id=request.POST.get('warehouse'))
        project = get_object_or_404(Project, id=request.POST.get('project'))
        quantity = Decimal(request.POST.get('quantity'))

        # Check if enough stock is available
        if material.current_stock < quantity:
            messages.error(request, f'Not enough stock! Available: {material.current_stock} {material.unit}')
            return redirect('myapp:materials_consumption')
        
        # Create the transaction
        transaction = MaterialTransaction.objects.create(
            material=material,
            warehouse=warehouse,
            project=project,
            transaction_type='consumption',
            quantity=-quantity,  # Negative quantity for consumption
            unit_price=material.unit_price,
            notes=request.POST.get('notes'),
        )
        transaction.log_action('CREATE', user=request.user.customuser, request=request)

        # Update material stock
        material.current_stock -= quantity
        material.save()
        material.log_action('UPDATE', user=request.user.customuser, request=request)

        messages.success(request, 'Material consumption recorded successfully.')
        return redirect('myapp:material_transaction')

    context = {
        'materials': Material.objects.all(),
        'warehouses': Warehouse.objects.all(),
        'projects': Project.objects.all(),
    }
    return render(request, 'myapp/record_consumption.html', context)

@login_required
def subcontractor_dashboard(request):
    """Dashboard view for subcontractor management."""
    # Get subcontractor statistics
    total_subcontractors = Subcontractor.objects.count()
    active_subcontractors = Subcontractor.objects.filter(status='ACTIVE').count()
    
    subcontractors_by_type = Subcontractor.objects.values(
        'specialization'
    ).annotate(
        count=Count('id')
    ).order_by('specialization')
    
    recent_subcontractors = Subcontractor.objects.order_by('-created_at')[:5]
    
    # Get project assignments through SubcontractorAssignment
    subcontractor_projects = Project.objects.filter(
        subcontractor_assignments__isnull=False
    ).distinct().count()

    context = {
        'total_subcontractors': total_subcontractors,
        'active_subcontractors': active_subcontractors,
        'subcontractors_by_type': subcontractors_by_type,
        'recent_subcontractors': recent_subcontractors,
        'subcontractor_projects': subcontractor_projects
    }
    return render(request, 'myapp/subcontractors/dashboard.html', context)

@login_required
def subcontractor_list(request):
    """List view for all subcontractors."""
    subcontractors = Subcontractor.objects.all().order_by('name')
    return render(request, 'myapp/subcontractors/list.html', {'subcontractors': subcontractors})

@login_required
def subcontractor_create(request):
    """Create a new subcontractor."""
    if request.method == 'POST':
        form = SubcontractorForm(request.POST)
        if form.is_valid():
            subcontractor = form.save()
            subcontractor.log_action('CREATE', user=request.user.customuser, request=request)
            messages.success(request, 'Subcontractor created successfully.')
            return redirect('myapp:subcontractor_list')
    else:
        form = SubcontractorForm()
    return render(request, 'myapp/subcontractors/form.html', {'form': form, 'action': 'Create'})

@login_required
def subcontractor_edit(request, pk):
    """Edit an existing subcontractor."""
    subcontractor = get_object_or_404(Subcontractor, pk=pk)
    if request.method == 'POST':
        form = SubcontractorForm(request.POST, instance=subcontractor)
        if form.is_valid():
            form.save()
            subcontractor.log_action('UPDATE', user=request.user.customuser, request=request)
            messages.success(request, 'Subcontractor updated successfully.')
            return redirect('myapp:subcontractor_list')
    else:
        form = SubcontractorForm(instance=subcontractor)
    return render(request, 'myapp/subcontractors/form.html', {'form': form, 'action': 'Edit'})

@login_required
def subcontractor_delete(request, pk):
    """Delete a subcontractor."""
    subcontractor = get_object_or_404(Subcontractor, pk=pk)
    if request.method == 'POST':
        subcontractor.log_action('DELETE', user=request.user.customuser, request=request)
        subcontractor.delete()
        messages.success(request, 'Subcontractor deleted successfully.')
        return redirect('myapp:subcontractor_list')
    return render(request, 'myapp/subcontractors/confirm_delete.html', {'subcontractor': subcontractor})

@login_required
def cost_dashboard(request):
    """Dashboard view for cost management."""
    # Get cost statistics
    total_material_cost = MaterialTransaction.objects.filter(
        transaction_type='DELIVERY'
    ).annotate(
        total=ExpressionWrapper(
            F('quantity') * F('unit_price'),
            output_field=DecimalField()
        )
    ).aggregate(
        total=Coalesce(Sum('total'), Value(0, output_field=DecimalField()))
    )['total']
    
    # Get costs by project
    project_costs = Project.objects.annotate(
        material_costs=Coalesce(
            Sum(
                F('material_transactions__quantity') * F('material_transactions__unit_price'),
                filter=Q(material_transactions__transaction_type='DELIVERY')
            ),
            Value(0, output_field=DecimalField())
        )
    ).order_by('-material_costs')[:5]
    
    # Get costs by material category
    costs_by_category = MaterialTransaction.objects.filter(
        transaction_type='DELIVERY'
    ).values(
        'material__category__name'
    ).annotate(
        total_cost=Sum(
            F('quantity') * F('unit_price'),
            output_field=DecimalField()
        )
    ).order_by('-total_cost')[:5]
    
    # Get recent transactions
    recent_transactions = MaterialTransaction.objects.select_related(
        'project', 'material'
    ).filter(
        transaction_type='DELIVERY'
    ).annotate(
        total_cost=ExpressionWrapper(
            F('quantity') * F('unit_price'),
            output_field=DecimalField()
        )
    ).order_by('-date')[:10]
    
    # Calculate monthly costs
    monthly_costs = MaterialTransaction.objects.filter(
        transaction_type='DELIVERY'
    ).annotate(
        month=TruncMonth('date'),
        transaction_cost=ExpressionWrapper(
            F('quantity') * F('unit_price'),
            output_field=DecimalField()
        )
    ).values('month').annotate(
        total=Sum('transaction_cost')
    ).order_by('-month')[:6]

    context = {
        'total_material_cost': total_material_cost,
        'project_costs': project_costs,
        'costs_by_category': costs_by_category,
        'recent_transactions': recent_transactions,
        'monthly_costs': monthly_costs
    }
    return render(request, 'myapp/costs/dashboard.html', context)

@login_required
def bom_list(request):
    """View for Bill of Materials list."""
    cost_centers = CostCenter.objects.prefetch_related('cost_items').all()
    return render(request, 'myapp/costs/bom_list.html', {'cost_centers': cost_centers})

@login_required
def cost_analysis(request):
    """View for cost analysis and reports with budget comparison."""
    try:
        # Get actual costs by type with totals
        cost_by_type = CostItem.objects.values('cost_type').annotate(
            total=Sum(F('quantity') * F('unit_price'))
        ).order_by('-total')
        
        # Get costs by month for trend analysis
        costs_by_month = CostItem.objects.annotate(
            month=TruncMonth('created_at')
        ).values('month').annotate(
            total=Sum(F('quantity') * F('unit_price'))
        ).order_by('month')
        
        # Get budgeted costs from planning phase
        budget_by_type = BudgetItem.objects.values('category').annotate(
            total=Sum('estimated_cost')
        ).order_by('-total')
        
        # Calculate totals
        total_actual_cost = sum(item['total'] for item in cost_by_type)
        total_budget = sum(item['total'] for item in budget_by_type)
        
        # Calculate monthly spending rate and forecast
        if costs_by_month:
            months = len(costs_by_month)
            if months > 0:
                monthly_rate = total_actual_cost / months
                year_forecast = monthly_rate * 12
            else:
                year_forecast = 0
        else:
            year_forecast = 0
        
        # Prepare data for charts
        cost_types = set()
        for item in chain(cost_by_type, budget_by_type):
            cost_type = item.get('cost_type') or item.get('category')
            if cost_type:
                cost_types.add(cost_type)
        
        chart_data = {
            'labels': list(cost_types),
            'actual_costs': [],
            'budgeted_costs': [],
            'trend_labels': [item['month'].strftime('%Y-%m') for item in costs_by_month],
            'trend_data': [float(item['total']) for item in costs_by_month]
        }
        
        for cost_type in cost_types:
            actual = next((item['total'] for item in cost_by_type if item['cost_type'] == cost_type), 0)
            budgeted = next((item['total'] for item in budget_by_type if item['category'] == cost_type), 0)
            chart_data['actual_costs'].append(float(actual))
            chart_data['budgeted_costs'].append(float(budgeted))
        
        # Add percentage and variance to cost analysis
        combined_costs = []
        for cost_type in cost_types:
            actual = next((item for item in cost_by_type if item['cost_type'] == cost_type), {'total': 0})
            budgeted = next((item for item in budget_by_type if item['category'] == cost_type), {'total': 0})
            variance = actual['total'] - budgeted['total']
            percentage = (actual['total'] / total_actual_cost * 100) if total_actual_cost > 0 else 0
            
            combined_costs.append({
                'cost_type': cost_type,
                'actual': actual['total'],
                'budgeted': budgeted['total'],
                'variance': variance,
                'percentage': percentage
            })
        
        # Get recent cost items
        cost_items = CostItem.objects.select_related('cost_center').order_by('-created_at')[:10]
        
        context = {
            'combined_costs': combined_costs,
            'cost_items': cost_items,
            'total_actual_cost': total_actual_cost,
            'total_budget': total_budget,
            'total_variance': total_actual_cost - total_budget,
            'year_forecast': year_forecast,
            'chart_data': chart_data
        }
        
        # Handle export request
        if request.GET.get('export') == 'csv':
            import csv
            from django.http import HttpResponse
            
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="cost_analysis.csv"'
            
            writer = csv.writer(response)
            writer.writerow(['Cost Type', 'Budgeted Amount', 'Actual Amount', 'Variance', '% of Total'])
            
            for item in combined_costs:
                writer.writerow([
                    item['cost_type'],
                    item['budgeted'],
                    item['actual'],
                    item['variance'],
                    f"{item['percentage']:.1f}%"
                ])
            
            writer.writerow([])
            writer.writerow(['Total', total_budget, total_actual_cost, total_actual_cost - total_budget, '100%'])
            
            return response
            
        return render(request, 'myapp/costs/analysis.html', context)
    except Exception as e:
        messages.error(request, f"Error in cost analysis: {str(e)}")
        return redirect('myapp:home')

# Maintenance Views
@login_required
def maintenance_dashboard(request):
    """Dashboard view for maintenance management."""
    # Get maintenance statistics
    total_maintenance = EquipmentMaintenance.objects.count()
    pending_maintenance = EquipmentMaintenance.objects.filter(completed_date__isnull=True).count()
    completed_maintenance = EquipmentMaintenance.objects.filter(completed_date__isnull=False).count()
    upcoming_maintenance = EquipmentMaintenance.objects.filter(
        scheduled_date__gte=timezone.now().date(),
        completed_date__isnull=True
    ).order_by('scheduled_date')[:5]
    
    # Get maintenance by type statistics
    maintenance_by_type = EquipmentMaintenance.objects.values('maintenance_type').annotate(
        count=Count('id')
    )
    
    # Calculate total maintenance cost
    total_cost = EquipmentMaintenance.objects.aggregate(
        total=Sum('cost')
    )['total'] or 0
    
    context = {
        'total_maintenance': total_maintenance,
        'pending_maintenance': pending_maintenance,
        'completed_maintenance': completed_maintenance,
        'upcoming_maintenance': upcoming_maintenance,
        'maintenance_by_type': maintenance_by_type,
        'total_cost': total_cost,
    }
    return render(request, 'myapp/maintenance_dashboard.html', context)

@login_required
def maintenance_list(request):
    """List view for maintenance records."""
    maintenance_records = EquipmentMaintenance.objects.all().order_by('-scheduled_date')
    return render(request, 'myapp/maintenance_list.html', {'maintenance_records': maintenance_records})

@login_required
def maintenance_create(request):
    """Create a new maintenance record."""
    if request.method == 'POST':
        form = MaintenanceForm(request.POST)
        if form.is_valid():
            maintenance = form.save(commit=False)
            maintenance.performed_by = request.user.customuser
            maintenance.save()
            maintenance.log_action('CREATE', user=request.user.customuser, request=request)
            messages.success(request, 'Maintenance record created successfully.')
            return redirect('myapp:maintenance_list')
    else:
        form = MaintenanceForm()
    
    return render(request, 'myapp/maintenance_form.html', {
        'form': form,
        'title': 'Create Maintenance Record'
    })

@login_required
def maintenance_edit(request, pk):
    """Edit an existing maintenance record."""
    maintenance = get_object_or_404(EquipmentMaintenance, pk=pk)
    if request.method == 'POST':
        form = MaintenanceForm(request.POST, instance=maintenance)
        if form.is_valid():
            form.save()
            maintenance.log_action('UPDATE', user=request.user.customuser, request=request)
            messages.success(request, 'Maintenance record updated successfully.')
            return redirect('myapp:maintenance_list')
    else:
        form = MaintenanceForm(instance=maintenance)
    
    return render(request, 'myapp/maintenance_form.html', {
        'form': form,
        'title': 'Edit Maintenance Record',
        'maintenance': maintenance
    })

@login_required
def maintenance_delete(request, pk):
    """Delete a maintenance record."""
    maintenance = get_object_or_404(EquipmentMaintenance, pk=pk)
    if request.method == 'POST':
        maintenance.log_action('DELETE', user=request.user.customuser, request=request)
        maintenance.delete()
        messages.success(request, 'Maintenance record deleted successfully.')
        return redirect('myapp:maintenance_list')
    return render(request, 'myapp/maintenance_confirm_delete.html', {'maintenance': maintenance})

@login_required
def maintenance_calendar(request):
    """Calendar view for maintenance schedules."""
    maintenance_records = EquipmentMaintenance.objects.all()
    return render(request, 'myapp/maintenance_calendar.html', {'maintenance_records': maintenance_records})

@login_required
def material_delivery(request):
    materials = Material.objects.all()
    warehouses = Warehouse.objects.filter(is_active=True)
    
    # Get filter parameters
    material_id = request.GET.get('material')
    warehouse_id = request.GET.get('warehouse')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    # Base queryset
    deliveries = MaterialTransaction.objects.filter(transaction_type='DELIVERY')
    
    # Apply filters
    if material_id:
        deliveries = deliveries.filter(material_id=material_id)
    if warehouse_id:
        deliveries = deliveries.filter(warehouse_id=warehouse_id)
    if date_from:
        deliveries = deliveries.filter(date__gte=date_from)
    if date_to:
        deliveries = deliveries.filter(date__lte=date_to)
        
    # Order by date
    deliveries = deliveries.order_by('-date', '-created_at')
    
    # Handle form submission
    if request.method == 'POST':
        form = MaterialTransactionForm(request.POST)
        if form.is_valid():
            transaction = form.save(commit=False)
            transaction.transaction_type = 'DELIVERY'
            transaction.created_by = request.user.customuser
            transaction.save()
            transaction.log_action('CREATE', user=request.user.customuser, request=request)
            
            # Update material stock
            material = transaction.material
            material.current_stock += transaction.quantity
            material.save()
            material.log_action('UPDATE', user=request.user.customuser, request=request)
            
            messages.success(request, 'Delivery recorded successfully.')
            return redirect('myapp:material_delivery')
    else:
        form = MaterialTransactionForm()
    
    # Pagination
    paginator = Paginator(deliveries, 10)
    page = request.GET.get('page')
    deliveries = paginator.get_page(page)
    
    context = {
        'deliveries': deliveries,
        'materials': materials,
        'warehouses': warehouses,
        'form': form,
        'selected_material': material_id,
        'selected_warehouse': warehouse_id,
        'date_from': date_from,
        'date_to': date_to,
    }
    
    return render(request, 'myapp/inventory/material_delivery.html', context)

@login_required
def material_consumption(request):
    materials = Material.objects.all()
    warehouses = Warehouse.objects.filter(is_active=True)
    projects = Project.objects.filter(status='ACTIVE')
    
    # Get filter parameters
    material_id = request.GET.get('material')
    warehouse_id = request.GET.get('warehouse')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    # Base queryset
    consumptions = MaterialTransaction.objects.filter(transaction_type='CONSUMPTION')
    
    # Apply filters
    if material_id:
        consumptions = consumptions.filter(material_id=material_id)
    if warehouse_id:
        consumptions = consumptions.filter(warehouse_id=warehouse_id)
    if date_from:
        consumptions = consumptions.filter(date__gte=date_from)
    if date_to:
        consumptions = consumptions.filter(date__lte=date_to)
        
    # Order by date
    consumptions = consumptions.order_by('-date', '-created_at')
    
    # Handle form submission
    if request.method == 'POST':
        form = MaterialTransactionForm(request.POST)
        if form.is_valid():
            transaction = form.save(commit=False)
            transaction.transaction_type = 'CONSUMPTION'
            transaction.created_by = request.user.customuser
            
            # Check if there's enough stock
            if transaction.quantity > transaction.material.current_stock:
                messages.error(request, 'Not enough stock available.')
                return redirect('myapp:material_consumption')
                
            transaction.save()
            transaction.log_action('CREATE', user=request.user.customuser, request=request)
            
            # Update material stock
            material = transaction.material
            material.current_stock -= transaction.quantity
            material.save()
            material.log_action('UPDATE', user=request.user.customuser, request=request)
            
            messages.success(request, 'Consumption recorded successfully.')
            return redirect('myapp:material_consumption')
    else:
        form = MaterialTransactionForm()
    
    # Pagination
    paginator = Paginator(consumptions, 10)
    page = request.GET.get('page')
    consumptions = paginator.get_page(page)
    
    context = {
        'consumptions': consumptions,
        'materials': materials,
        'warehouses': warehouses,
        'projects': projects,
        'form': form,
        'selected_material': material_id,
        'selected_warehouse': warehouse_id,
        'date_from': date_from,
        'date_to': date_to,
    }
    
    return render(request, 'myapp/inventory/material_consumption.html', context)

@login_required
def material_transaction_list(request):
    """View for listing all material transactions."""
    transactions = MaterialTransaction.objects.select_related(
        'material', 'project'
    ).order_by('-date')
    
    # Handle search
    search_query = request.GET.get('search')
    if search_query:
        transactions = transactions.filter(
            Q(material__name__icontains=search_query) |
            Q(project__name__icontains=search_query)
        )
    
    # Handle date filtering
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date:
        transactions = transactions.filter(date__gte=start_date)
    if end_date:
        transactions = transactions.filter(date__lte=end_date)
    
    # Handle transaction type filtering
    transaction_type = request.GET.get('type')
    if transaction_type:
        transactions = transactions.filter(transaction_type=transaction_type)
    
    # Pagination
    paginator = Paginator(transactions, 25)  # Show 25 transactions per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'start_date': start_date,
        'end_date': end_date,
        'transaction_type': transaction_type,
    }
    return render(request, 'myapp/material_transaction_list.html', context)

@login_required
def milestone_list(request, project_id):
    project = get_object_or_404(Project, pk=project_id)
    milestones = project.milestones.all().order_by('due_date')
    return render(request, 'myapp/milestones/milestone_list.html', {
        'project': project,
        'milestones': milestones
    })

@login_required
def add_milestone(request, project_id):
    project = get_object_or_404(Project, pk=project_id)
    
    if request.method == 'POST':
        form = MilestoneForm(request.POST)
        if form.is_valid():
            milestone = form.save(commit=False)
            milestone.project = project
            milestone.save()
            milestone.log_action('CREATE', user=request.user.customuser, request=request)
            messages.success(request, 'Milestone added successfully.')
            return redirect('myapp:milestone_list', project_id=project.id)
    else:
        form = MilestoneForm()
    
    return render(request, 'myapp/milestones/milestone_form.html', {
        'form': form,
        'project': project
    })

@login_required
def mark_milestone_completed(request, milestone_id):
    milestone = get_object_or_404(Milestone, pk=milestone_id)
    if request.method == 'POST':
        milestone.mark_as_completed()
        milestone.log_action('UPDATE', user=request.user.customuser, request=request)
        messages.success(request, f'Milestone "{milestone.name}" marked as completed.')
    return redirect('myapp:milestone_list', project_id=milestone.project.id)

# Project Planning Views
@login_required
def project_planning(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    
    # Get planning phase tasks
    planning_tasks = project.tasks.filter(phase='PLANNING').order_by('start_date')
    
    # Get budget items with totals
    budget_items = project.budget_items.all()
    total_estimated = sum(item.estimated_cost for item in budget_items)
    total_actual = sum(item.actual_cost or 0 for item in budget_items)
    
    # Get resource allocations
    resource_allocations = project.resource_allocations.all()
    
    # Calculate task dependencies
    tasks_with_dependencies = []
    for task in planning_tasks:
        dependencies = task.get_prerequisite_tasks()
        tasks_with_dependencies.append({
            'task': task,
            'dependencies': dependencies
        })
    
    context = {
        'project': project,
        'planning_tasks': planning_tasks,
        'budget_items': budget_items,
        'total_estimated': total_estimated,
        'total_actual': total_actual,
        'resource_allocations': resource_allocations,
        'tasks_with_dependencies': tasks_with_dependencies,
    }
    return render(request, 'myapp/planning/project_planning.html', context)

@login_required
def budget_item_create(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    
    if request.method == 'POST':
        form = BudgetItemForm(request.POST)
        if form.is_valid():
            budget_item = form.save(commit=False)
            budget_item.project = project
            budget_item.save()
            budget_item.log_action('CREATE', user=request.user.customuser, request=request)
            messages.success(request, 'Budget item added successfully.')
            return redirect('myapp:project_planning', project_id=project.id)
    else:
        form = BudgetItemForm()
    
    return render(request, 'myapp/planning/budget_item_form.html', {
        'form': form,
        'project': project
    })

@login_required
def resource_allocation_create(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    
    if request.method == 'POST':
        form = ResourceAllocationForm(request.POST)
        if form.is_valid():
            allocation = form.save(commit=False)
            allocation.project = project
            allocation.save()
            allocation.log_action('CREATE', user=request.user.customuser, request=request)
            messages.success(request, 'Resource allocation added successfully.')
            return redirect('myapp:project_planning', project_id=project.id)
    else:
        form = ResourceAllocationForm()
    
    return render(request, 'myapp/planning/resource_allocation_form.html', {
        'form': form,
        'project': project
    })

@login_required
def task_dependency_create(request, project_id, task_id):
    project = get_object_or_404(Project, id=project_id)
    task = get_object_or_404(Task, id=task_id, project=project)
    
    if request.method == 'POST':
        form = TaskDependencyForm(request.POST, task=task, project=project)
        if form.is_valid():
            dependency = form.save(commit=False)
            dependency.task = task
            dependency.save()
            dependency.log_action('CREATE', user=request.user.customuser, request=request)
            messages.success(request, 'Task dependency added successfully.')
            return redirect('myapp:project_planning', project_id=project.id)
    else:
        form = TaskDependencyForm(task=task, project=project)
    
    return render(request, 'myapp/planning/task_dependency_form.html', {
        'form': form,
        'project': project,
        'task': task
    })

@login_required
def planning_task_create(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    
    if request.method == 'POST':
        form = TaskForm(request.POST, project=project)
        if form.is_valid():
            task = form.save(commit=False)
            task.project = project
            task.created_by = request.user
            task.phase = 'PLANNING'
            task.save()
            task.log_action('CREATE', user=request.user.customuser, request=request)
            messages.success(request, 'Planning task added successfully.')
            return redirect('myapp:project_planning', project_id=project.id)
    else:
        form = TaskForm(project=project)
    
    return render(request, 'myapp/planning/planning_task_form.html', {
        'form': form,
        'project': project
    })

@login_required
def start_project_execution(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    
    if request.method == 'POST':
        # Validate that all necessary planning is complete
        planning_tasks = project.tasks.filter(phase='PLANNING', status='NOT_STARTED').exists()
        budget_items = project.budget_items.exists()
        resource_allocations = project.resource_allocations.exists()
        
        if planning_tasks or not budget_items or not resource_allocations:
            messages.error(request, 'Please complete all planning tasks, budget items, and resource allocations before starting execution.')
            return redirect('myapp:project_planning', project_id=project.id)
        
        # Update project status to ACTIVE
        project.status = 'ACTIVE'
        project.save()
        project.log_action('UPDATE', user=request.user.customuser, request=request)
        
        messages.success(request, 'Project has been moved to execution phase.')
        return redirect('myapp:project_detail', project_id=project.id)
    
    return render(request, 'myapp/planning/start_execution_confirm.html', {
        'project': project
    })

@login_required
@admin_required
def change_user_role(request, user_id):
    """View for changing a user's role"""
    if request.method == 'POST':
        user = get_object_or_404(CustomUser, id=user_id)
        new_role = request.POST.get('role')
        
        from .roles import assign_role_to_user
        if assign_role_to_user(user.user, new_role):
            messages.success(request, f"Role updated to {new_role} for {user.user.username}")
        else:
            messages.error(request, "Failed to update role")
            
    return redirect('myapp:user_management')

@login_required
@admin_required
def toggle_user_active(request, user_id):
    """View for toggling user active status"""
    if request.method == 'POST':
        user = get_object_or_404(CustomUser, id=user_id)
        user.user.is_active = not user.user.is_active
        user.user.save()
        
        status = "activated" if user.user.is_active else "deactivated"
        messages.success(request, f"User {user.user.username} has been {status}")
            
    return redirect('myapp:user_management')

@login_required
@role_required(['Admin', 'Project Manager', 'Worker', 'Viewer'])
def task_detail(request, pk):
    """View for displaying task details"""
    task = get_object_or_404(Task, pk=pk)
    user = request.user.customuser
    
    # Workers can only view their assigned tasks
    if user.has_role('Worker') and task.assigned_to != user:
        raise PermissionDenied
    
    context = {
        'task': task,
        'can_edit': user.has_role('Admin') or user.has_role('Project Manager') or 
                   (user.has_role('Worker') and task.assigned_to == user),
        'can_delete': user.has_role('Admin') or user.has_role('Project Manager')
    }
    return render(request, 'myapp/task_detail.html', context)

@login_required
def audit_log_list(request):
    """View for listing audit logs with filtering and pagination."""
    # Debug information
    print("User:", request.user)
    print("Is authenticated:", request.user.is_authenticated)
    if hasattr(request.user, 'customuser'):
        print("Role:", request.user.customuser.get_role())
    
    queryset = AuditLog.objects.select_related('user').all()
    
    # Apply filters
    user_id = request.GET.get('user')
    action = request.GET.get('action')
    model = request.GET.get('model')
    date_range = request.GET.get('date_range')
    
    if user_id:
        queryset = queryset.filter(user_id=user_id)
    if action:
        queryset = queryset.filter(action=action)
    if model:
        queryset = queryset.filter(model_name=model)
    if date_range:
        try:
            start_date, end_date = date_range.split(' - ')
            start_date = datetime.strptime(start_date, '%m/%d/%Y')
            end_date = datetime.strptime(end_date, '%m/%d/%Y') + timedelta(days=1)
            queryset = queryset.filter(action_time__range=(start_date, end_date))
        except (ValueError, IndexError):
            pass
    
    # Pagination
    paginator = Paginator(queryset, 50)  # Show 50 logs per page
    page = request.GET.get('page')
    audit_logs = paginator.get_page(page)
    
    # Get filter options
    users = CustomUser.objects.filter(
        id__in=AuditLog.objects.values_list('user', flat=True).distinct()
    )
    models = AuditLog.objects.values_list('model_name', flat=True).distinct()
    
    context = {
        'audit_logs': audit_logs,
        'users': users,
        'actions': AuditLog.ACTION_CHOICES,
        'models': models,
        'selected_user': user_id,
        'selected_action': action,
        'selected_model': model,
        'date_range': date_range,
    }
    
    return render(request, 'myapp/audit_log_list.html', context)
